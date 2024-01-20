import os, sys,io, json
import json
import firebase_admin
from firebase_admin import credentials, storage
from datetime import datetime, timedelta
from dotenv import load_dotenv
from telethon.sync import TelegramClient
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
import pandas as pd

# Load environment variables
DIR = os.getcwd()
active_users_json_path = os.path.join(DIR,"MarketUtils", "active_users.json")
broker_filepath = os.path.join(DIR,"MarketUtils", "broker.json")
env_path = os.path.join(DIR, '.env')
session_filepath = os.path.join(DIR,'MarketUtils', 'Telegram','+918618221715.session')

load_dotenv(env_path)
api_id = os.getenv('telethon_api_id')
api_hash = os.getenv('telethon_api_hash')

sys.path.append(DIR)


# Retrieve values from .env for Firebase and Telegram
firebase_credentials_path = os.getenv('FIREBASE_CREDENTIALS_PATH')
database_url = os.getenv('DATABASE_URL')
storage_bucket = os.getenv('STORAGE_BUCKET')
api_id = os.getenv('TELETHON_API_ID')
api_hash = os.getenv('TELETHON_API_HASH')

# # Initialize Firebase app if it hasn't been initialized yet
# if not firebase_admin._apps:
#     cred = credentials.Certificate(firebase_credentials_path)
#     firebase_admin.initialize_app(cred, {
#         'databaseURL': database_url,
#         'storageBucket': storage_bucket
#     })

# Function to find the start date of the current complete week
def get_current_week_range():
    """Finds and returns the start date of the current complete week."""
    today = datetime.now()
    start_date = today - timedelta(days=today.weekday())
    end_date = start_date + timedelta(days=4)
    return start_date, end_date

# # Function to get the free cash/margin available for a user
# def get_cashmargin_value(user_data):
#     """Retrieves the cash margin available for a user based on their broker."""
#     active_users = general_calc.read_json_file(active_users_json_path)
#     for user in active_users:
#         if user['account_name'] == user_data['account_name']:
#             if user['broker'] == "aliceblue":
#                 cash_margin = cash_margin_available(user)
#             elif user['broker'] == "zerodha":
#                 cash_margin = cash_balance(user)
#             try:
#                 return float(cash_margin)  # Ensure cash_margin is a float
#             except ValueError:
#                 print(f"Invalid cash margin value for {user['account_name']}: {cash_margin}")
#                 return 0.0  # Return a default value or handle as appropriate
#     return 0.0  # If user or broker not found

# Function to send a message via Telegram
def send_telegram_message(phone_number, message):
    """Sends a message to a specified phone number via Telegram."""
    with TelegramClient(session_filepath, api_id, api_hash) as client:
        client.send_message(phone_number, message, parse_mode='md') 

# # Function to generate a formatted message for weekly reports
# def generate_message(user, excel_file_name, net_pnl, cash_margin_value, trademan_account_value, trademan_invested, drawdown, commission, actual_account_value, difference_value, start_date, end_date):
#     """Generates and returns a formatted weekly report message."""
#     message = f"Weekly Summary for {user['account_name']} ({start_date.strftime('%B %d')} to {end_date.strftime('%B %d')})\n\n"
    
#     # Process the details separately
#     details = process_DTD(excel_file_name)
#     for detail in details:
#         message += f"{detail}\n"

#     message += f"\n**Net PnL: {custom_format(net_pnl)}**\n\n"
#     message += f"Free Cash: {custom_format(cash_margin_value)}\n"
#     message += f"Trademan Invested: {custom_format(trademan_invested)}\n"
#     message += f"Estimated Account Value: {custom_format(trademan_account_value)}\n"
#     message += f"Actual Account Value: {custom_format(actual_account_value)}\n"
#     message += f"Difference: {custom_format(difference_value)}\n\n"

#     # Only add the commission to the message if it's greater than 0
#     if commission > 0:
#         message += f"Commission: {custom_format(commission)}\n\n"

#     # Add the drawdown to the message if base_capital is less than actual_account_value
#     if drawdown < 0:
#         message += f"Drawdown: -{custom_format(abs(drawdown))}\n\n"

#     message += "Best regards,\n**Serendipity Trading Firm**"
#     return message

# Function to calculate net PnL
# def calculate_net_pnl(excel_file_name):
#     # Load the Excel file within the function
#     df_dtd, _ = load_excel(excel_file_name)  # Only load the 'DTD' sheet

#     net_pnl = 0.0  # Initialize net PnL

#     if df_dtd is not None:
#         # Ensure 'Amount' column is numeric
#         df_dtd['Amount'] = pd.to_numeric(df_dtd['Amount'].replace('[₹,]', '', regex=True).replace('-', '-0'), errors='coerce').fillna(0)

#         # Filter data for the current week
#         start_week, end_week = get_current_week()
#         current_week_df = df_dtd[(df_dtd['Date'].dt.date >= start_week) & (df_dtd['Date'].dt.date <= end_week)]

#         # Calculate net PnL as the sum of 'Amount' column
#         net_pnl = current_week_df['Amount'].sum()
#         return net_pnl

# # Function to calculate Trademan invested value (customize this function according to your logic)
# def calculate_trademan_invested(excel_file_name):
#     # Load the 'Holdings' sheet
#     _, df_holdings = load_excel(excel_file_name)

#     total_margin_used = 0.0  # Initialize the net PnL for holdings without an exit time

#     if df_holdings is not None:
#         # Clean the column names
#         df_holdings.columns = df_holdings.columns.str.strip().str.title()

#         # Ensure 'Net_Pnl' and 'Exit_Time' columns exist and are in the correct format
#         if 'Net_Pnl' in df_holdings.columns and 'Exit_Time' in df_holdings.columns:
#             # Convert 'Net_Pnl' to numeric, replacing non-numeric with 0
#             df_holdings['Net_Pnl'] = pd.to_numeric(df_holdings['Net_Pnl'].replace('[₹,]', '', regex=True), errors='coerce').fillna(0)

#             # Filter rows where 'Exit_Time' is NaN (i.e., the trade is still active)
#             active_holdings = df_holdings[df_holdings['Exit_Time'].isna()]

#             # Sum 'Net_Pnl' for these active holdings
#             total_margin_used = active_holdings['Net_Pnl'].sum()

#     # Return the total margin used (or modify as needed for your specific use case)
#     return  total_margin_used

# Function to read base capital from basecapital.txt
def read_base_capital(file_path):
    base_capital = {}
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            for line in lines:
                parts = line.strip().split(':')
                if len(parts) == 2:
                    user = parts[0].strip()
                    balance_str = parts[1].strip()  # Directly strip whitespace
                    base_capital[user] = float(balance_str)  # Convert the balance string to a float
    except FileNotFoundError:
        print("basecapital.txt not found.")
    except ValueError as e:
        print(f"Error parsing base capital: {e}")
    return base_capital

# Function to calculate commission and drawdown
def calculate_commission_and_drawdown(user, actual_account_value, base_capital):
    user_name = user['account_name']
    commission = 0.0
    drawdown = 0.0

    if user_name in base_capital:
        user_base_capital = base_capital[user_name]
        
        if actual_account_value > user_base_capital:
            # Commission is positive when there's a profit
            commission = actual_account_value - user_base_capital
        elif actual_account_value < user_base_capital:
            # Drawdown is negative when there's a loss
            drawdown = actual_account_value - user_base_capital
        # print(f"Account: {user_name}, Commission: {commission}, Drawdown: {drawdown}")
    else:
        print(f"Base capital not found for {user_name}.")

    return commission, drawdown

# Main function to execute the script for generating weekly reports
def main():
    # csv_file_path = r'C:\Users\user\OneDrive\Desktop\TradeManV1\SampleData\Sample_Kite_ledger.csv'
    # # Example usage
    # kite_categorized_dfs = process_kite_ledger(csv_file_path)
    # kite_net_values = calculate_kite_net_values(kite_categorized_dfs)

    # # Print the net values
    # print(kite_net_values)
    
    # Example usage
    excel_file_path = r'C:\Users\user\OneDrive\Desktop\TradeManV1\SampleData\ledger\alice\Sample_alice_ledger.xlsx'  # Replace with your Excel file path
    process_alice_ledger(excel_file_path)
    # alice_net_values = calculate_alice_net_values(categorized_dfs_v2)


    # Print the net values
    # print(alice_net_values)  
                    

def process_kite_ledger(csv_file_path):
    # Define patterns for categorizing transactions
    patterns = {
        "Deposits": ["Funds added using UPI", "Opening Balance", "Funds added using payment gateway from YY0222"],
        "Withdrawals": ["Funds transferred back as part of quarterly settlement", "Payout of"],
        "Charges": ["Being payment gateway charges debited", "DPCharges", "Reversal of Brokerage", "Kite Connect API Charges", "Call and Trade charges", "AMC for Demat Account", "DP Charges for Sale of"],
        "Trades": ["Span margin blocked for NSE F&O", "Exposure margin blocked for NSE F&O", "Span margin reversed for NSE F&O", "Exposure margin reversed for NSE F&O", "Net obligation for Equity F&O", "Net settlement for Equity", "Net obligation for Currency F&O"]
    }

    # Load the CSV file
    ledger_data = pd.read_csv(csv_file_path)

    # Function to categorize a transaction
    def categorize_transaction(particulars):
        for category, patterns_list in patterns.items():
            for pattern in patterns_list:
                if pattern in particulars:
                    return category
        return "Other"

    # Categorize each transaction
    ledger_data['Category'] = ledger_data['particulars'].apply(categorize_transaction)

    # Filter out transactions with 'Closing Balance'
    ledger_data_filtered = ledger_data[ledger_data['particulars'] != 'Closing Balance']

    # Create dataframes for each category
    categorized_dfs = {category: ledger_data_filtered[ledger_data_filtered['Category'] == category] for category in patterns.keys()}
    
    #save categorized_dfs to csv as {category}.csv
    for category, df in categorized_dfs.items():
        df.to_csv(f"{category}.csv")

    return categorized_dfs

def calculate_kite_net_values(categorized_dfs):
    # Calculate net values for each category
    net_values = {category: df['debit'].sum() - df['credit'].sum() for category, df in categorized_dfs.items()}
    return net_values

def process_alice_ledger(excel_file_path):
    # Load the workbook and access the worksheet
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # List to store rows
    rows = []

    # Get a list of all merged cells ranges
    merged_ranges = [mr.bounds for mr in sheet.merged_cells.ranges]

    # Iterate over the rows starting from the 5th row
    for row in sheet.iter_rows(min_row=5):
        row_data = []
        for cell in row:
            merged_range = next(((start_row, start_col, end_row, end_col) for start_row, start_col, end_row, end_col in merged_ranges
                                if cell.row in range(start_row, end_row+1) and cell.column in range(start_col, end_col+1)), None)
            if merged_range:
                if (cell.row, cell.column) == (merged_range[0], merged_range[1]):
                    row_data.append(cell.value)
                continue
            else:
                row_data.append(cell.value)
        rows.append(row_data[:9])
        
    print(rows)
        
    # Create a DataFrame from the rows
    filtered_data = pd.DataFrame(rows, columns=['Date', 'Voucher', 'VoucherNo', 'Code', 'Narration', 'ChqNo', 'Debit', 'Credit', 'Running Bal'])

    # Filtering out rows where any values are NaN across columns A:K (indexes 0 to 10)
    filtered_data = filtered_data.dropna(how='any', subset=filtered_data.columns[:9])
    print(filtered_data)

    # Filtering out header rows by identifying rows that match the header pattern
    headers = ['Date', 'Voucher', 'VoucherNo', 'Code', 'Narration', 'ChqNo', 'Debit', 'Credit', 'Running Bal']
    filtered_data = filtered_data[~filtered_data[0].astype(str).str.contains('Date')]

    # Combine 'VoucherNo' and 'Running Bal' columns which are split across two columns
    filtered_data['VoucherNo'] = filtered_data[2].astype(str) + filtered_data[3].astype(str)
    filtered_data['Running Bal'] = filtered_data[9].astype(str) + filtered_data[10].astype(str)

    # Drop the original split columns from the dataframe
    filtered_data.drop(columns=[2, 3, 9, 10], inplace=True)
    
    # Assigning proper header to the filtered data
    filtered_data.columns = headers + list(filtered_data.columns[len(headers):])

    # Define patterns for categorization with updated rules
    patterns = {
        "NetDeposits": [
            "PAYMENT DONE VIA : RAZORPAY NET",
            "RECEIVED AMOUNT THROUGH HDFC-CMS(OTH)",
            "PAYMENT DONE VIA : RAZORPAY UPI"
        ],
        "NetWithdrawals": [
            "PAYOUT OF FUNDS"
        ],
        "Netcharges": [
            "CGST", "SGST", "BENEFICIARY CHARGES",
            "DP MAINTENANCE CHARGES FOR THE PERIOD",
            "CALL AND TRADE OR SQUARE OFF CHARGES FOR",
            "BENEFICIARY CHARGES FOR SETT NO",
            "BEING PAYMENT GATEWAY CHARGES DEBITED -"
        ],
        "trades": [
            "BILL ENTRY FOR FO-", "BILL ENTRY FOR M-","BILL ENTRY FOR Z-"
        ],
        "ignore": [
            "INTER EXCHANGE SETL JV FROM NSEFNO TO BSECASH",
            "INTER EXCHANGE SETL JV FROM BSECASH TO NSEFNO",
            "Narration"
        ]
    }

    # Apply categorization patterns
    def categorize_transaction(narration):
        if pd.isna(narration):
            return 'ignore'
        for category, pattern_list in patterns.items():
            for pattern in pattern_list:
                if pattern in narration:
                    return category
        return "Other"

    # Apply categorization patterns
    filtered_data['Category'] = filtered_data.apply(
        lambda row: categorize_transaction(row['Narration']), axis=1)

    # Recalculate the net values for each category after re-categorization
    categorized_dfs_final = {
        category: filtered_data[filtered_data['Category'] == category]
        for category in patterns if category != "ignore"
}

    # # Convert 'Debit' and 'Credit' columns to numeric and fill NaNs with 0
    # for df in categorized_dfs_final.values():
    #     df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
    #     df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)

    # Calculate net values for each category
    # net_values_final = {category: df['Debit'].sum() - df['Credit'].sum() for category, df in categorized_dfs_final.items()}

    # Save each categorized dataframe to a CSV file
    for category, df in categorized_dfs_final.items():
        print(f'Saving {category} transactions to CSV...')
        df.to_csv(f'{category}_transactions.csv', index=False)

    # Check for any 'Other' transactions left and save to CSV
    other_transactions_final = filtered_data[filtered_data['Category'] == 'Other']
    other_transactions_final.to_csv('Other_transactions_final.csv', index=False)

    return categorized_dfs_final, other_transactions_final

def calculate_alice_net_values(categorized_dfs):
    # Calculate net values for each category
    net_values = {category: df['Debit'].sum() - df['Credit'].sum() for category, df in categorized_dfs.items()}
    return net_values

if __name__ == "__main__":
    main()


